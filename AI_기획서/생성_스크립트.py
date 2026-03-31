# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from copy import copy

wb = openpyxl.Workbook()

# === 명시적 RGB 색상 (테마 색상 사용하지 않음) ===
fill_dark = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')   # #000000 순검정
fill_gray = PatternFill(start_color='FF262626', end_color='FF262626', fill_type='solid')   # #262626 진회색

font_white = 'FFFFFFFF'
def mf(bold=False, size=11.0):
    return Font(name='맑은 고딕', size=size, bold=bold, color=font_white)

ft_toc = mf(True, 14.0)
ft_h1 = mf(True, 15.0)
ft_h1b = mf(True, 14.0)
ft_h2 = mf(True, 13.0)
ft_n = mf(False, 11.0)
ft_bn = mf(True, 11.0)
va = Alignment(vertical='center')

# ============================================================
# 시트1: 체력 회복 시스템
# ============================================================
ws = wb.active
ws.title = '체력 회복 시스템'
ws.sheet_view.showGridLines = False
ws.sheet_format.defaultColWidth = 2.625
ws.sheet_format.defaultRowHeight = 16.5
ws.column_dimensions['A'].width = 2.625
ws.column_dimensions['C'].width = 2.625
ws.column_dimensions['D'].width = 2.625

# === 전체 셀에 #262626 fill 적용 (엑셀 화면에서 잘리지 않도록 넉넉하게) ===
MAX_ROW = 200
MAX_COL = 50
for r in range(1, MAX_ROW + 1):
    for col in range(1, MAX_COL + 1):
        x = ws.cell(row=r, column=col)
        x.fill = copy(fill_gray)
        x.font = ft_n
        x.alignment = va

def c(row, col, val, font=ft_n):
    x = ws.cell(row=row, column=col, value=val if val else None)
    x.font = font
    # fill은 이미 전체 적용됨, 덮어쓸 필요 없음

def dark_row(row):
    """큰 제목 행: A~H 전체를 #000000으로 채움"""
    for col in range(1, MAX_COL + 1):
        ws.cell(row=row, column=col).fill = copy(fill_dark)

def h(row, ht):
    ws.row_dimensions[row].height = ht

# ==============================
# 목차
# ==============================
dark_row(2)
c(2, 2, '목차', ft_toc)
h(2, 20.25)

c(3, 2, '', ft_h1b)
h(3, 20.25)

c(4, 2, '1. 개요', ft_n)
c(5, 3, '(1) 체력 회복이란', ft_n)
c(6, 3, '(2) 회복 수단 종류', ft_n)

c(8, 2, '2. 회복 규칙', ft_n)
c(9, 3, '(1) 즉시 회복', ft_n)
c(10, 3, '(2) 지속 회복', ft_n)

# ==============================
# 1. 개요
# ==============================
dark_row(13)
c(13, 2, '1. 개요', ft_h1)
h(13, 24.0)

c(14, 2, '', ft_h1b)
h(14, 20.25)

c(15, 3, '(1) 체력 회복이란', ft_h2)
h(15, 19.5)

c(16, 4, 'A. 전투 중 줄어든 체력을 되돌리는 모든 행위를 뜻한다', ft_n)
c(17, 5, '1) 아이템 사용, 스킬 발동, 환경 상호작용 등이 포함된다', ft_n)
c(18, 5, '2) 회복량은 최대 체력을 넘지 않는다', ft_n)

c(20, 3, '(2) 회복 수단 종류', ft_h2)
h(20, 19.5)

c(21, 4, 'A. 즉시 회복', ft_n)
c(22, 5, '1) 사용하면 곧바로 고정된 양만큼 체력을 돌려받는다', ft_n)
c(23, 6, 'a. 회복 아이템 (포션 등)', ft_n)
c(24, 6, 'b. 동료 힐 스킬', ft_n)

c(26, 4, 'B. 지속 회복', ft_n)
c(27, 5, '1) 일정 시간 동안 조금씩 나눠서 체력을 돌려받는다', ft_n)
c(28, 6, 'a. 회복 음식 (서서히 회복)', ft_n)
c(29, 6, 'b. 휴식 장소 (모닥불, 웨이포인트)', ft_n)

# ==============================
# 2. 회복 규칙
# ==============================
dark_row(33)
c(33, 2, '2. 회복 규칙', ft_h1b)
h(33, 20.25)

c(34, 2, '', ft_h1b)
h(34, 20.25)

c(35, 3, '(1) 즉시 회복', ft_h2)
h(35, 19.5)

c(36, 4, 'A. 기본 규칙', ft_n)
c(37, 5, '1) 아이템 사용 시 DT_Item의 HealAmount만큼 즉시 회복된다', ft_n)
c(38, 5, '2) 전투 중에도 사용할 수 있다', ft_n)
c(39, 6, 'a. 사용 모션 도중 피격 시 회복이 취소된다', ft_n)
c(40, 6, 'b. 쿨타임은 아이템마다 따로 적용된다', ft_n)

c(42, 4, 'B. 동료 힐 스킬', ft_n)
c(43, 5, '1) AI 동료가 리더 체력이 30% 이하일 때 자동으로 사용한다', ft_n)
c(44, 6, 'a. 동료 스킬 쿨타임이 적용된다', ft_n)
c(45, 6, 'b. 동료가 사망 중이면 사용할 수 없다', ft_n)

c(47, 3, '(2) 지속 회복', ft_h2)
h(47, 19.5)

c(48, 4, 'A. 틱 회복 규칙', ft_n)
c(49, 5, '1) 1초 간격으로 HealPerTick만큼 회복된다', ft_n)
c(50, 5, '2) 피격을 받아도 지속 회복은 끊기지 않는다', ft_n)
c(51, 5, '3) 최대 체력에 도달하면 남은 회복은 사라진다', ft_n)

c(53, 4, 'B. 휴식 회복', ft_n)
c(54, 5, '1) 모닥불이나 웨이포인트에서 휴식하면 체력을 전부 돌려받는다', ft_n)
c(55, 6, 'a. 휴식을 시작하려면 먼저 전투 상태가 풀려야 한다', ft_n)
c(56, 7, '- 근처에 적이 있으면 휴식을 시작할 수 없다', ft_n)
c(57, 8, '· 주변에 위협이 남아 있습니다 메시지가 뜬다', ft_n)

# ============================================================
# 시트2: 디자인 구성 정보
# ============================================================
ws2 = wb.create_sheet('디자인 구성 정보')
ws2.sheet_view.showGridLines = False
ws2.sheet_format.defaultColWidth = 2.625
ws2.sheet_format.defaultRowHeight = 16.5
ws2.column_dimensions['A'].width = 2.625

for r in range(1, 10):
    for col in range(1, 9):
        x = ws2.cell(row=r, column=col)
        x.fill = copy(fill_gray)
        x.font = ft_n
        x.alignment = va

ws2.cell(row=3, column=2, value='1. 중요한 내용은 bold 처리를 한다.').font = ft_n
ws2.cell(row=4, column=2, value='2. 더 중요한 내용은 주황색 컬러로 텍스트를 칠한다.').font = ft_n
ws2.cell(row=5, column=2, value='3. 가장 중요한 내용은 bold하고 주황색 컬러로 텍스트를 칠한다.').font = ft_n

wb.save('샘플_체력회복시스템.xlsx')
print('완료!')

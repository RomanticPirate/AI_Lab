# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import zipfile, os

OUTPUT = '복장_상태_변경_규칙.xlsx'

# === 색상 ===
fill_dark = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
fill_gray = PatternFill(start_color='FF262626', end_color='FF262626', fill_type='solid')
fill_tbl_h = PatternFill(start_color='FF404040', end_color='FF404040', fill_type='solid')

fw = 'FFFFFFFF'
def mf(bold=False, size=11.0):
    return Font(name='맑은 고딕', size=size, bold=bold, color=fw)
def mf_orange(bold=False, size=11.0):
    return Font(name='맑은 고딕', size=size, bold=bold, color='FFFF8C00')

ft_toc = mf(True, 14.0)
ft_h1 = mf(True, 15.0)
ft_h1b = mf(True, 14.0)
ft_h2 = mf(True, 13.0)
ft_n = mf(False, 11.0)
ft_bn = mf(True, 11.0)
ft_bo = mf_orange(True, 11.0)
va = Alignment(vertical='center')
va_c = Alignment(vertical='center', horizontal='center')

thin_gray = Side(style='thin', color='FF808080')
tbl_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

MAX_ROW = 1000
MAX_COL = 108

wb = openpyxl.Workbook()

# ============================================================
# 시트1: 복장 상태 변경 규칙
# ============================================================
ws = wb.active
ws.title = '복장 상태 변경 규칙'
ws.sheet_view.showGridLines = False
ws.sheet_format.defaultColWidth = 2.625
ws.sheet_format.defaultRowHeight = 16.5

print('시트1 배경 채우는 중...')
for r in range(1, MAX_ROW + 1):
    for col in range(1, MAX_COL + 1):
        x = ws.cell(row=r, column=col)
        x.fill = fill_gray
        x.font = ft_n
        x.alignment = va

def c(row, col, val, font=ft_n):
    x = ws.cell(row=row, column=col, value=val if val else None)
    x.font = font

def dark_row(row):
    for col in range(1, MAX_COL + 1):
        ws.cell(row=row, column=col).fill = fill_dark

def h(row, ht):
    ws.row_dimensions[row].height = ht

def tbl(row, cols, values, font=ft_n, header=False):
    al = va_c if header else va
    for (sc, ec), val in zip(cols, values):
        if ec > sc:
            ws.merge_cells(start_row=row, start_column=sc, end_row=row, end_column=ec)
        cell = ws.cell(row=row, column=sc, value=val)
        cell.font = font
        cell.alignment = al
        for cc in range(sc, ec + 1):
            ws.cell(row=row, column=cc).border = tbl_border
            if header:
                ws.cell(row=row, column=cc).fill = fill_tbl_h

# ==============================
# 목차
# ==============================
dark_row(2); c(2, 2, '목차', ft_toc); h(2, 20.25)
c(3, 2, '', ft_h1b); h(3, 20.25)

c(4, 2, '1. 개요', ft_n)
c(5, 3, '(1) 시스템 목적/의도', ft_n)
c(6, 3, '(2) 핵심 요약', ft_n)
c(7, 3, '(3) 핵심 용어 정의', ft_n)

c(9, 2, '2. 복장 데미지 시스템', ft_n)
c(10, 3, '(1) 기본 규칙', ft_n)
c(11, 3, '(2) 전체 흐름 순서도', ft_n)
c(12, 3, '(3) 진행 규칙', ft_n)
c(13, 3, '(4) 진행 예시', ft_n)
c(14, 3, '(5) 데이터 구조', ft_n)
c(15, 3, '(6) 원상 복구 규칙', ft_n)
c(16, 3, '(7) 저장 / 로드 규칙', ft_n)

c(18, 2, '3. 복장 젖음 시스템', ft_n)
c(19, 3, '(1) 기본 규칙', ft_n)
c(20, 3, '(2) 전체 흐름 순서도', ft_n)
c(21, 3, '(3) 적용 규칙', ft_n)
c(22, 3, '(4) 진행 예시', ft_n)
c(23, 3, '(5) 데이터 운영', ft_n)
c(24, 3, '(6) 원상 복구 규칙', ft_n)

c(26, 2, '4. 예외처리 / 엣지케이스', ft_n)
c(28, 2, '5. 설정 변수 표', ft_n)

# ==============================
# 1. 개요
# ==============================
dark_row(31); c(31, 2, '1. 개요', ft_h1); h(31, 24.0)
c(32, 2, '', ft_h1b); h(32, 20.25)

c(33, 3, '(1) 시스템 목적/의도', ft_h2); h(33, 19.5)
c(34, 4, 'A. 복장 상태 변경 시스템은 전투와 환경 상호작용이 캐릭터 외형에 시각적으로 반영되는 시스템이다', ft_n)
c(35, 4, 'B. 전투의 격렬함과 환경 접촉의 현실감을 플레이어가 직접 눈으로 확인할 수 있게 한다', ft_n)
c(36, 4, 'C. 단순한 비주얼 연출이 아니라, 게임 내 상황이 캐릭터에 남는 경험을 제공하는 것이 목적이다', ft_bn)

c(38, 3, '(2) 핵심 요약', ft_h2); h(38, 19.5)
c(39, 4, 'A. 복장 데미지: 피격 시 CostumeDamagePoint가 누적되어 단계별 외형이 반영된다', ft_bo)
c(40, 4, 'B. 복장 젖음: Water Volume 접촉 부위에 젖음이 적용되고, 이탈 후 서서히 마른다', ft_bo)

c(42, 3, '(3) 핵심 용어 정의', ft_h2); h(42, 19.5)
T_TERM = [(5, 14), (15, 35)]
tbl(43, T_TERM, ['용어', '설명'], ft_bn, header=True)
for i, (term, desc) in enumerate([
    ('CostumeDamagePoint', '피격 시 복장 데미지에 부여되는 점수 (StatusEffect별 설정)'),
    ('복장 데미지 단계', '누적 점수가 기준값에 도달할 때 변경되는 외형 단계'),
    ('StatusEffect', '피격 반응을 정의하는 데이터 Row'),
    ('Water Volume', '물 접촉을 감지하는 영역'),
    ('Blend 값', '젖음/마름의 시각적 전환 속도 (Fade In/Out 공용)'),
    ('마름 타이머', 'Water Volume 이탈 후 마르기까지의 남은 시간'),
    ('DT_CutScene.Damage', '컷씬 시작 시 복장 데미지 유지(Keep) / 초기화(Reset) 제어'),
]):
    tbl(44 + i, T_TERM, [term, desc])

# ==============================
# 2. 복장 데미지 시스템
# ==============================
dark_row(53); c(53, 2, '2. 복장 데미지 시스템', ft_h1); h(53, 24.0)
c(54, 2, '', ft_h1b); h(54, 20.25)

# (1) 기본 규칙
c(55, 3, '(1) 기본 규칙', ft_h2); h(55, 19.5)
c(56, 4, 'A. 리더와 동료가 피격당할 때 복장 데미지가 진행된다', ft_n)
c(57, 4, 'B. 피격 시 CostumeDamagePoint가 쌓이고, 기준값에 도달하면 다음 단계 외형이 반영된다', ft_bo)
c(58, 4, 'C. 누적 점수가 기준에 도달하지 못하면 점수만 오르고 외형은 바뀌지 않는다', ft_n)
c(59, 4, 'D. 복장 데미지는 원상 복구 이벤트가 발생하기 전까지 유지된다', ft_n)
c(60, 4, 'E. 복장 데미지 상태는 세이브 / 로드 대상이다', ft_bn)

# (2) 전체 흐름 순서도
c(62, 3, '(2) 전체 흐름 순서도', ft_h2); h(62, 19.5)
c(63, 4, 'A. 아래 순서도 참고', ft_n)
# 순서도 1: 0-indexed rows 63~111 (Excel R64~R112)
# "추가 피격 발생?" 다이아몬드 추가로 6행 확장

# (3) 진행 규칙
c(114, 3, '(3) 진행 규칙', ft_h2); h(114, 19.5)
c(115, 4, 'A. 전투 중 피격이 발생하면 발동한 StatusEffect를 확인한다', ft_n)
c(116, 4, 'B. 해당 StatusEffect의 CostumeDamagePoint를 현재 누적 점수에 더한다', ft_n)
c(117, 4, 'C. 누적 점수가 다음 단계 기준에 도달했는지 확인한다', ft_bn)
c(118, 5, '1) 미도달: 점수만 갱신하고 현재 외형을 유지한다', ft_n)
c(119, 5, '2) 도달: 복장 데미지 단계를 올리고 다음 단계 외형을 반영한다', ft_bo)
c(120, 4, 'D. 이후 다시 피격이 발생하면 같은 규칙을 반복한다', ft_n)

# (4) 진행 예시
c(122, 3, '(4) 진행 예시', ft_h2); h(122, 19.5)
c(123, 4, 'A. 복장 데미지 진행 예시', ft_bn)
T2 = [(5,7), (8,14), (15,17), (18,20), (21,23), (24,26), (27,32)]
tbl(124, T2, ['피격', 'StatusEffect', '획득', '누적', '기준값', '도달', '결과'], ft_bn, header=True)
for i, vals in enumerate([
    ('1차', 'Stagger', '+1', '1', '5', '미도달', '변화 없음'),
    ('2차', 'Knockdown', '+2', '3', '5', '미도달', '변화 없음'),
    ('3차', 'Grabbed', '+2', '5', '5', '도달', '단계 1 반영'),
    ('4차', 'Stagger', '+1', '6', '8', '미도달', '변화 없음'),
    ('5차', 'Break Attacked', '+2', '8', '8', '도달', '단계 2 반영'),
]):
    font = ft_bo if vals[5] == '도달' else ft_n
    tbl(125 + i, T2, vals, font)

c(131, 5, '1) 원상 복구 트리거 발생 시 → 복장 기본 상태로 복구', ft_n)

# (5) 데이터 구조
c(133, 3, '(5) 데이터 구조', ft_h2); h(133, 19.5)
c(134, 4, 'A. DT_StatusEffect', ft_bn)
c(135, 5, '1) 컬럼 추가: CostumeDamagePoint', ft_n)
c(136, 5, '2) 해당 StatusEffect가 복장 데미지에 부여하는 점수', ft_n)
c(137, 5, '3) 피격 시 발동한 StatusEffect의 CostumeDamagePoint를 현재 점수에 누적한다', ft_n)

c(139, 4, 'B. DT_CutScene', ft_bn)
c(140, 5, '1) 컬럼 추가: Damage (Enum)', ft_n)
c(141, 5, '2) Keep: 컷씬 시작 시 현재 복장 데미지 상태를 유지한다', ft_n)
c(142, 5, '3) Reset: 컷씬 시작 시 복장을 기본 상태로 되돌린다', ft_n)

c(144, 4, 'C. DT_StatusEffect 데이터 예시', ft_bn)
T1 = [(5,9), (10,15), (16,23), (24,35)]
tbl(145, T1, ['Row', 'Type', 'CostumeDamagePoint', '설명'], ft_bn, header=True)
for i, vals in enumerate([
    ('Stagger_Lv0', 'Stagger', '1', '경미한 피격'),
    ('Knockdown_Lv1', 'Knockdown', '2', '넘어짐을 유발하는 피격'),
    ('BreakAttacked', 'Break Attacked', '2', '강한 반응을 일으키는 피격'),
    ('Dead', 'Dead', '0', '복장 데미지 부여 대상 아님'),
]):
    tbl(146 + i, T1, vals)

# (6) 원상 복구 규칙
c(152, 3, '(6) 원상 복구 규칙', ft_h2); h(152, 19.5)
c(153, 4, 'A. 복장 데미지 단계와 외형을 기본 상태로 되돌리는 처리', ft_n)
c(154, 4, 'B. 원상 복구 트리거', ft_bo)
c(155, 5, '1) 복장 교체', ft_n)
c(156, 5, '2) 지역 이동 (로딩이 발생하는 모든 순간 이동)', ft_n)
c(157, 5, '3) 웨이포인트 휴식', ft_n)
c(158, 5, '4) 컷씬 시작 (DT_CutScene.Damage = Reset인 경우만)', ft_n)
c(160, 4, 'C. 컷씬 Reset 사용 의도', ft_n)
c(161, 5, '1) 복장 데미지가 유지되면 연출 분위기를 해칠 수 있다', ft_n)
c(162, 5, '2) 중요 컷씬만 Reset 적용', ft_n)
c(163, 5, '3) 짧은 진입 컷씬이나 경량 시퀀스는 Keep 적용 가능', ft_n)

# (7) 저장 / 로드 규칙
c(165, 3, '(7) 저장 / 로드 규칙', ft_h2); h(165, 19.5)
c(166, 4, 'A. 복장 데미지 상태는 저장 대상이다', ft_bo)
c(167, 4, 'B. 세이브 시 현재 복장 데미지 단계를 저장한다', ft_n)
c(168, 4, 'C. 로드 시 저장된 복장 데미지 단계를 복원한다', ft_n)

# ==============================
# 3. 복장 젖음 시스템
# ==============================
dark_row(171); c(171, 2, '3. 복장 젖음 시스템', ft_h1); h(171, 24.0)
c(172, 2, '', ft_h1b); h(172, 20.25)

# (1) 기본 규칙
c(173, 3, '(1) 기본 규칙', ft_h2); h(173, 19.5)
c(174, 4, 'A. Water Volume과 접촉한 부위에만 젖음이 적용된다', ft_n)
c(175, 4, 'B. 젖음은 부위 단위로 관리한다', ft_bn)
c(176, 4, 'C. 같은 부위가 다시 젖으면 마름 타이머를 초기화하고 전체 유지 시간부터 다시 진행한다', ft_bo)
c(177, 4, 'D. 다른 부위가 젖는 경우 각 부위는 독립적으로 젖음/마름을 진행한다', ft_n)
c(178, 4, 'E. 젖는 속도와 마르는 속도에 같은 Blend 값을 쓴다', ft_n)
c(179, 4, 'F. 컷씬 중에도 젖음 유지, 갱신, 마름이 진행된다', ft_n)
c(180, 4, 'G. 젖음 상태는 세이브 / 로드 대상이 아니다', ft_bo)

# (2) 전체 흐름 순서도
c(182, 3, '(2) 전체 흐름 순서도', ft_h2); h(182, 19.5)
c(183, 4, 'A. 아래 순서도 참고', ft_n)
# 순서도 2: 0-indexed rows 183~225 (Excel R184~R226)

# (3) 적용 규칙
c(228, 3, '(3) 적용 규칙', ft_h2); h(228, 19.5)
c(229, 4, 'A. Water Volume 접촉 시', ft_bn)
c(230, 5, '1) 접촉한 부위에 즉시 젖음이 시작된다', ft_n)
c(231, 5, '2) 부위별 Blend 값에 따라 서서히 젖음이 반영된다', ft_n)
c(233, 4, 'B. Water Volume 이탈 시', ft_bn)
c(234, 5, '1) 해당 부위의 마름 타이머가 시작된다', ft_n)
c(236, 4, 'C. 같은 부위 재접촉 시', ft_bn)
c(237, 5, '1) 해당 부위의 마름 타이머를 초기화한다', ft_n)
c(238, 5, '2) 전체 유지 시간부터 다시 진행한다', ft_n)
c(240, 4, 'D. 마름 시간 도달 시', ft_bn)
c(241, 5, '1) 부위별 Blend 값에 따라 서서히 마른다', ft_n)
c(242, 5, '2) 완전히 마르면 해당 부위 젖음을 해제한다', ft_n)

# (4) 진행 예시
c(244, 3, '(4) 진행 예시', ft_h2); h(244, 19.5)
c(245, 4, 'A. 젖음 진행 예시 타임라인', ft_bn)
T3 = [(5,7), (8,13), (14,19), (20,25), (26,31), (32,37), (38,43)]
tbl(246, T3, ['부위', 'T0', 'T1', 'T2', 'T3', 'T4', 'T5'], ft_bn, header=True)
for i, vals in enumerate([
    ('A', '젖음 시작', '젖음 유지', '이탈→마름 시작', '마름 진행', '마름 완료', '건조 유지'),
    ('B', '젖음 시작', '젖음 유지', '이탈→마름 시작', '재접촉→초기화', '이탈→새 타이머', '마름 완료'),
    ('C', '건조', '건조', '건조', '젖음 시작', '젖음 유지', '이탈→마름'),
]):
    tbl(247 + i, T3, vals)

c(251, 5, '1) B는 T2에 마름 타이머가 시작되지만 T3에 다시 젖음이 발생한다', ft_n)
c(252, 5, '2) B의 마름 타이머는 초기화되고, 새 유지 시간부터 다시 진행한다', ft_n)

# (5) 데이터 운영
c(254, 3, '(5) 데이터 운영', ft_h2); h(254, 19.5)
c(255, 4, 'A. 젖음 상태는 부위 단위로 개별 관리한다', ft_n)
c(256, 4, 'B. 부위별로 필요한 정보', ft_bn)
c(257, 5, '1) 현재 젖음 여부', ft_n)
c(258, 5, '2) 마름 타이머 진행 여부', ft_n)
c(259, 5, '3) 남은 마름 시간', ft_n)
c(260, 4, 'C. 부위를 몇 개로 나눌지는 프로그램/그래픽 파트와 협의한다', ft_n)

# (6) 원상 복구 규칙
c(262, 3, '(6) 원상 복구 규칙', ft_h2); h(262, 19.5)
c(263, 4, 'A. 젖은 부위가 모두 마름 완료 상태로 돌아가는 처리', ft_n)
c(264, 4, 'B. 부위별 개별 처리', ft_bn)
c(265, 5, '1) 같은 부위 재접촉 시 해당 부위만 타이머를 초기화한다', ft_n)
c(266, 5, '2) 다른 부위는 기존 타이머를 유지한다', ft_n)
c(267, 4, 'C. 세이브 / 로드', ft_bn)
c(268, 5, '1) 젖음 상태는 저장하지 않는다', ft_bo)
c(269, 5, '2) 로드 후 기본 건조 상태로 시작한다', ft_n)

# ==============================
# 4. 예외처리 / 엣지케이스
# ==============================
dark_row(272); c(272, 2, '4. 예외처리 / 엣지케이스', ft_h1); h(272, 24.0)
c(273, 2, '', ft_h1b); h(273, 20.25)

c(274, 3, '(1) 복장 데미지와 젖음 동시 발생', ft_h2); h(274, 19.5)
c(275, 4, 'A. 두 시스템은 완전히 독립적으로 처리한다', ft_bn)
c(276, 4, 'B. 데미지 상태와 젖음 상태는 서로 영향을 주지 않는다', ft_n)
c(277, 4, 'C. 외형은 데미지 단계와 젖음 상태가 동시에 반영된다', ft_n)

c(279, 3, '(2) 최대 단계 도달 후 추가 피격', ft_h2); h(279, 19.5)
c(280, 4, 'A. 점수는 계속 누적되지만 외형은 최대 단계를 유지한다', ft_bn)
c(281, 4, 'B. 원상 복구 트리거 발생 시 기본 상태로 돌아간다', ft_n)

c(283, 3, '(3) 컷씬 중 복합 상태 처리', ft_h2); h(283, 19.5)
c(284, 4, 'A. 젖음은 컷씬 중에도 유지/갱신/마름이 계속 진행된다', ft_n)
c(285, 4, 'B. 복장 데미지는 DT_CutScene.Damage 설정에 따라 결정된다', ft_bo)
c(286, 5, '1) Keep: 현재 데미지 상태 유지', ft_n)
c(287, 5, '2) Reset: 기본 상태로 복구', ft_n)

c(289, 3, '(4) 로드 시 복합 상태 복원', ft_h2); h(289, 19.5)
c(290, 4, 'A. 복장 데미지는 저장된 단계로 복원된다', ft_bn)
c(291, 4, 'B. 젖음은 저장 대상이 아니므로 건조 상태로 시작한다', ft_bn)

# ==============================
# 5. 설정 변수 표
# ==============================
dark_row(294); c(294, 2, '5. 설정 변수 표', ft_h1b); h(294, 20.25)
c(295, 2, '', ft_h1b); h(295, 20.25)

T4 = [(4,12), (13,28), (29,37)]
tbl(296, T4, ['항목', '설명', '설정 위치'], ft_bn, header=True)
for i, vals in enumerate([
    ('CostumeDamagePoint', '피격 시 복장 데미지에 부여하는 점수', 'DT_StatusEffect'),
    ('복장 데미지 기준값', '다음 외형 단계로 넘어가는 기준 점수', '복장 데미지 단계 데이터'),
    ('젖음 유지 시간', 'Water Volume 이탈 후 마르기까지 유지되는 시간', '젖음 전역 설정값'),
    ('젖음 Blend 값', '젖는 속도 / 마르는 속도 공용 전환 시간', '젖음 전역 설정값'),
    ('Damage', '컷씬 시작 시 복장 데미지 Keep / Reset 제어', 'DT_CutScene'),
]):
    tbl(297 + i, T4, vals)

# ============================================================
# 시트2: 레퍼런스
# ============================================================
ws2 = wb.create_sheet('레퍼런스')
ws2.sheet_view.showGridLines = False
ws2.sheet_format.defaultColWidth = 2.625
ws2.sheet_format.defaultRowHeight = 16.5

print('시트2 배경 채우는 중...')
for rr in range(1, MAX_ROW + 1):
    for col in range(1, MAX_COL + 1):
        x = ws2.cell(row=rr, column=col)
        x.fill = fill_gray
        x.font = ft_n
        x.alignment = va

def c2(row, col, val, font=ft_n):
    x = ws2.cell(row=row, column=col, value=val if val else None)
    x.font = font

def dark_row2(row):
    for col in range(1, MAX_COL + 1):
        ws2.cell(row=row, column=col).fill = fill_dark

dark_row2(2)
c2(2, 2, '레퍼런스', ft_toc)

refs = [
    ('Far Cry 6', [
        '젖음을 정적/동적 두 축으로 분리 (Ubisoft 공식 기술 문서)',
        '캐릭터, 무기, 차량은 dynamic wetness 대상',
        '접촉 시 시작 → 노출 끊기면 서서히 마름 → 전환이 즉시 끊기지 않음',
    ], 'https://www.ubisoft.com/en-us/company/how-we-make-games/technology/articles/simulating-tropical-weather-in-far-cry-6'),
    ('Ghost of Yōtei', [
        '캐릭터가 wet/bloody/muddy/snowy 상태로 레이어링됨 (PlayStation 기술 딥다이브)',
        '부위별 상태 관리 구조는 다른 상태 표현 확장에도 재사용 가능',
    ], 'https://blog.playstation.com/2025/10/23/ghost-of-yotei-tech-deep-dive/'),
    ('The Last of Us Part II', [
        '움직이는 캐릭터 표면에 water drips, blood, snow 동적 효과 (GDC/SIGGRAPH)',
        '젖음이 컷씬 중에도 유지/변화/건조 진행된다는 정책을 정당화하는 근거',
    ], 'https://www.gdcvault.com/play/1027356/Enhancement-of-Particle-Simulation-Using'),
    ('Red Dead Redemption 2', [
        '진흙탕에 넘어지면 옷이 젖고, 사냥한 동물의 피가 재킷에 묻고, 탄흔이 남음',
        '복장 데미지를 전투와 연동된 누적형 비주얼 변화로 두는 방향의 근거',
    ], 'https://www.digipen.edu/showcase/news/digipen-grad-andy-kibler-puts-the-clothes-on-the-cowboy-in-red-dead-redemption-2'),
    ("Baldur's Gate 3", [
        '캐릭터가 dirty/sweaty/bloodied/bruised 상태가 누적, 특정 행위로 해제',
        '복장 데미지는 저장 대상, 젖음은 미저장 대상이라는 기획 방향에 참고',
    ], 'https://www.gamesradar.com/baldurs-gate-3-patch-lets-your-characters-get-literally-battered-and-bruised/'),
    ("Marvel's Spider-Man 2", [
        '전투 반복 시 suit damage가 scuff → 찢어짐까지 단계적으로 심해짐',
        '누적 점수와 상태 변경 기준을 별도로 두는 구조가 관리하기 쉬움',
    ], 'https://playerassist.com/does-marvels-spider-man-2-have-suit-damage-answered/'),
    ("Dragon's Dogma 계열", [
        'Drenched 상태는 물 접촉/강/비 등 환경 노출로 발생',
        '현재 범위에서는 외형 상태 표현까지만 확정, 전투 효과 연계는 추후 별도 검토',
    ], 'https://dragonsdogma2.wiki.fextralife.com/Drenched'),
]

r = 4
for idx, (title, points, url) in enumerate(refs):
    c2(r, 3, f'({idx + 1}) {title}', ft_h2)
    ws2.row_dimensions[r].height = 19.5
    r += 1
    for pi, p in enumerate(points):
        c2(r, 4, f'{"ABCDEFG"[pi]}. {p}', ft_n)
        r += 1
    c2(r, 4, f'출처: {url}', mf(False, 10.0))
    r += 2

# ============================================================
# 저장
# ============================================================
print('저장 중...')
wb.save(OUTPUT)
print(f'{OUTPUT} 생성 완료!')
